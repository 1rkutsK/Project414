from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes,
    filters, CallbackQueryHandler, ConversationHandler
)
from openpyxl import load_workbook
from apscheduler.schedulers.background import BackgroundScheduler
import datetime
import calendar
import os
import requests
import json
import pytz

TOKEN = "ТОКЕН"
WEATHER_API_KEY = "ТОКЕН"  # Нужно будет заменить на реальный ключ
CITY_NAME = "Moscow"  # Город по умолчанию
ADMINS = [АЙДИ]
DEFAULT_NOTIFICATION_TIME = 20  # Время рассылки по умолчанию (20:00)

# Состояния для ConversationHandler
WAITING_FOR_NOTE = 1
WAITING_FOR_FILE = 2
WAITING_FOR_SEARCH = 3
WAITING_FOR_DELETE = 4
WAITING_FOR_TIME = 5

user_groups = {}
subscribed_users = set()
user_notification_times = {}  # Хранение пользовательских настроек времени рассылки
schedule_data = []
available_groups = []
current_day_selections = {}
current_week_offsets = {}
user_notes = {}  # Структура: {user_id: {date_str: note}}
user_calendar_date = {}  # Для хранения выбранной даты в календаре


def load_schedule():
    global schedule_data, available_groups
    schedule_data = []
    available_groups = []
    directions = set()

    if not os.path.exists("schedule.xlsx"):
        return

    wb = load_workbook("schedule.xlsx")
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        group = row[0]
        # Пропускаем строки с пустым значением группы
        if not group:
            continue

        day = row[1]
        time = row[2]
        type_ = row[3]
        subject = row[4]
        room = row[5]
        teacher = row[6] if len(row) > 6 else ""
        subgroup = row[7] if len(row) > 7 else None
        week_num = row[8] if len(row) > 8 else None

        schedule_data.append({
            'group': group,
            'day': day,
            'time': time,
            'type': type_,
            'subject': subject,
            'room': str(room).strip(),
            'teacher': str(teacher).strip() if teacher else "",
            'subgroup': str(subgroup).strip() if subgroup else None,
            'week_num': int(week_num) if week_num in (1, 2) else None
        })

        directions.add(group[:3])

    available_groups = sorted(set(item['group'] for item in schedule_data if item['group']))
    return sorted(directions)


def get_current_week(date):
    # Начальная дата (26 мая 2024) - первая неделя
    start_date = datetime.datetime(2024, 5, 26)
    delta = date - start_date
    if delta.days < 0:
        # Для дат до начальной даты считаем от предыдущей недели
        delta_weeks = abs(delta.days) // 7 + 1
        return 1 if delta_weeks % 2 == 0 else 2
    else:
        # Для дат после начальной даты
        delta_weeks = delta.days // 7
        return 1 if delta_weeks % 2 == 0 else 2


def get_schedule_by_date(group, date, week_offset=0):
    current_week = get_current_week(date)

    weekday = date.strftime("%A")
    date_str = date.strftime("%d %B %Y").replace("January", "января").replace("February", "февраля") \
        .replace("March", "марта").replace("April", "апреля").replace("May", "мая") \
        .replace("June", "июня").replace("July", "июля").replace("August", "августа") \
        .replace("September", "сентября").replace("October", "октября") \
        .replace("November", "ноября").replace("December", "декабря")
    russian_days = {
        "Monday": "Понедельник", "Tuesday": "Вторник", "Wednesday": "Среда",
        "Thursday": "Четверг", "Friday": "Пятница", "Saturday": "Суббота",
        "Sunday": "Воскресенье"
    }
    day_rus = russian_days.get(weekday, "")

    lessons = [
        item for item in schedule_data
        if item['group'] == group
           and item['day'] == day_rus
           and (item['week_num'] is None or item['week_num'] == current_week)
    ]

    if not lessons:
        return f"\U0001F4C5 {date_str} {day_rus}\n\nЗанятий нет."

    grouped_lessons = {}
    for lesson in lessons:
        key = (lesson['time'], lesson['subject'], lesson['type'])
        if key not in grouped_lessons:
            grouped_lessons[key] = []
        grouped_lessons[key].append(lesson)

    lessons_sorted = sorted(grouped_lessons.items(),
                            key=lambda x: datetime.datetime.strptime(x[0][0].split("-")[0], "%H:%M"))

    week_type = "Вторая" if current_week == 2 else "Первая"
    result = f"\U0001F4C5 {date_str} {day_rus} ({week_type} неделя)\n\n"

    for (time, subject, type_), sub_lessons in lessons_sorted:
        pair_num = get_pair_number(time)
        result += f"<b>{pair_num} пара, {time}</b>\n[{type_}]\n<b>{subject}</b>\n"
        if len(sub_lessons) > 1 or sub_lessons[0]['subgroup']:
            for sub in sub_lessons:
                subgroup_info = f" (п/гр {sub['subgroup']})" if sub['subgroup'] else ""
                result += f"├─ Аудитория: {sub['room']}\n├─ Преподаватель: {sub['teacher']}{subgroup_info}\n"
        else:
            result += f"Аудитория: {sub_lessons[0]['room']}\nПреподаватель: {sub_lessons[0]['teacher']}\n"
        result += "\n"

    return result


def get_pair_number(time_str):
    time_ranges = [
        ("8:30-9:50", 1), ("10:05-11:25", 2), ("11:40-13:00", 3),
        ("13:45-15:05", 4), ("15:20-16:40", 5), ("16:55-18:15", 6),
        ("18:30-19:50", 7), ("20:00-21:20", 8)
    ]
    return next((num for t, num in time_ranges if t == time_str), 1)


def create_calendar_keyboard(year: int, month: int):
    keyboard = []

    # Заголовок с месяцем и годом
    month_names = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    ]
    keyboard.append([InlineKeyboardButton(
        f"{month_names[month - 1]} {year}",
        callback_data="ignore"
    )])

    # Дни недели
    days_of_week = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    keyboard.append([InlineKeyboardButton(day, callback_data="ignore") for day in days_of_week])

    # Получаем календарь на месяц
    cal = calendar.monthcalendar(year, month)

    # Добавляем дни
    for week in cal:
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(" ", callback_data="ignore"))
            else:
                row.append(InlineKeyboardButton(
                    str(day),
                    callback_data=f"calendar_{year}_{month}_{day}"
                ))
        keyboard.append(row)

    # Добавляем кнопки навигации
    nav_row = []
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    nav_row.extend([
        InlineKeyboardButton("◀️", callback_data=f"month_{prev_year}_{prev_month}"),
        InlineKeyboardButton("Отмена", callback_data="cancel_calendar"),
        InlineKeyboardButton("▶️", callback_data=f"month_{next_year}_{next_month}")
    ])
    keyboard.append(nav_row)

    return InlineKeyboardMarkup(keyboard)


def create_admin_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📤 Загрузить расписание", callback_data="upload_schedule")],
        [InlineKeyboardButton("◀ Назад", callback_data="go_back_to_main")]
    ])


async def show_admin_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id not in ADMINS:
        await update.message.reply_text("У вас нет доступа к админ-панели.")
        return

    await update.message.reply_text(
        "Панель администратора:",
        reply_markup=create_admin_keyboard()
    )


def create_main_menu_keyboard(user_id):
    keyboard = []
    is_subscribed = user_id in subscribed_users
    subscription_text = "📬 Рассылка: Вкл" if is_subscribed else "📭 Рассылка: Выкл"
    notification_time = get_user_notification_time(user_id)

    if user_id in ADMINS:
        keyboard.append([InlineKeyboardButton("👑 Админ-панель", callback_data="admin_panel")])

    keyboard.extend([
        [InlineKeyboardButton("📅 Посмотреть расписание", callback_data="view_schedule")],
        [InlineKeyboardButton("📝 Добавить заметку", callback_data="show_calendar")],
        [InlineKeyboardButton("🗂 Мои заметки", callback_data="notes_menu")],
        [InlineKeyboardButton("❓ Помощник", callback_data="ai_assistant")],
        [InlineKeyboardButton(subscription_text, callback_data="toggle_subscription")],
        [InlineKeyboardButton(f"⏰ Время рассылки: {notification_time:02d}:00", callback_data="set_time")],
        [InlineKeyboardButton("👥 Сменить группу", callback_data="change_group")]
    ])

    return InlineKeyboardMarkup(keyboard)


def create_notes_menu_keyboard():
    keyboard = [
        [InlineKeyboardButton("🔍 Поиск по тексту", callback_data="search_text")],
        [InlineKeyboardButton("📅 Поиск по дате", callback_data="search_date")],
        [InlineKeyboardButton("📝 Показать все заметки", callback_data="show_all")],
        [InlineKeyboardButton("🗑 Удалить заметку", callback_data="delete_note")],
        [InlineKeyboardButton("◀ Назад", callback_data="go_back_to_main")]
    ]
    return InlineKeyboardMarkup(keyboard)


def get_course_from_group(group: str) -> int:
    # Находим первую цифру в названии группы
    for char in group:
        if char.isdigit():
            return int(char)
    return 0


def create_course_keyboard():
    if not available_groups:
        return None

    # Получаем все доступные курсы
    courses = sorted(set(get_course_from_group(group) for group in available_groups))

    keyboard = []
    for course in courses:
        if course > 0:  # Пропускаем некорректные значения
            keyboard.append([InlineKeyboardButton(f"{course} курс", callback_data=f"course_{course}")])

    if len(keyboard) == 0:
        return None

    return InlineKeyboardMarkup(keyboard)


def create_direction_keyboard_for_course(course: int):
    # Получаем направления только для выбранного курса
    directions = set()
    for group in available_groups:
        if get_course_from_group(group) == course:
            directions.add(group[:3])

    keyboard = []
    for direction in sorted(directions):
        keyboard.append([InlineKeyboardButton(direction, callback_data=f"initial_direction_{direction}")])

    keyboard.append([InlineKeyboardButton("◀ К выбору курса", callback_data="back_to_courses")])
    return InlineKeyboardMarkup(keyboard)


def create_groups_keyboard_for_direction(direction: str, course: int):
    # Получаем группы только для выбранного направления и курса
    groups = sorted([
        group for group in available_groups
        if group.startswith(direction) and get_course_from_group(group) == course
    ])

    buttons = [[InlineKeyboardButton(group, callback_data=f"initial_group_{group}")] for group in groups]
    buttons.append([InlineKeyboardButton("◀ К направлениям", callback_data=f"back_to_directions_{course}")])
    return InlineKeyboardMarkup(buttons)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    current_week_offsets[user_id] = 0

    # Загружаем расписание, если оно еще не загружено
    if not available_groups:
        if user_id in ADMINS:
            await update.message.reply_text(
                "Расписание ещё не загружено. Загрузите его через админ-панель:",
                reply_markup=create_admin_keyboard()
            )
        else:
            await update.message.reply_text(
                "Расписание ещё не загружено. Обратитесь к администратору."
            )
        return

    # Показываем выбор курса
    keyboard = create_course_keyboard()
    if keyboard:
        await update.message.reply_text(
            "👋 Добро пожаловать! Для начала выберите курс:",
            reply_markup=keyboard
        )
    else:
        await update.message.reply_text(
            "❌ Ошибка: нет доступных групп. Обратитесь к администратору."
        )


async def request_schedule_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.edit_message_text(
        "Пожалуйста, отправьте файл расписания в формате Excel (.xlsx)",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀ Отмена", callback_data="cancel_upload")
        ]])
    )
    return WAITING_FOR_FILE


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if user_id not in ADMINS:
        await update.message.reply_text(
            "У вас нет прав загружать расписание.",
            reply_markup=create_main_menu_keyboard(user_id)
        )
        return

    document = update.message.document
    if not document.file_name.endswith(".xlsx"):
        await update.message.reply_text(
            "Нужен файл формата .xlsx",
            reply_markup=create_main_menu_keyboard(user_id)
        )
        return

    file = await context.bot.get_file(document.file_id)
    await file.download_to_drive("schedule.xlsx")
    load_schedule()

    await update.message.reply_text(
        "Расписание успешно загружено!",
        reply_markup=create_main_menu_keyboard(user_id)
    )


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    text = update.message.text.strip()

    # Проверяем, похож ли текст на запрос к AI
    if ("когда" in text.lower() and "следующ" in text.lower()) or \
            ("сколько" in text.lower() and "пар" in text.lower() and "недел" in text.lower()):
        await handle_ai_request(update, context)
        return

    # Проверяем, ожидаем ли мы дату для удаления заметки
    if context.user_data.get('waiting_for_delete'):
        del context.user_data['waiting_for_delete']
        try:
            date_obj = datetime.datetime.strptime(text, "%d.%m.%Y")
            if delete_note(user_id, date_obj):
                await update.message.reply_text(
                    f"✅ Заметка на {text} успешно удалена!",
                    reply_markup=create_notes_menu_keyboard()
                )
            else:
                await update.message.reply_text(
                    f"❌ Заметка на {text} не найдена!",
                    reply_markup=create_notes_menu_keyboard()
                )
        except ValueError:
            await update.message.reply_text(
                "⚠️ Неверный формат даты. Используйте формат ДД.ММ.ГГГГ, например: 25.03.2024",
                reply_markup=create_notes_menu_keyboard()
            )
        return

    # Проверяем, ожидаем ли мы поисковый запрос
    if context.user_data.get('waiting_for_search'):
        search_type = context.user_data.get('search_type')
        del context.user_data['waiting_for_search']

        if search_type == "search_date":
            try:
                date_obj = datetime.datetime.strptime(text, "%d.%m.%Y")
                search_date = date_obj.strftime("%Y-%m-%d")
                result = search_notes(user_id, date_str=search_date)
            except ValueError:
                await update.message.reply_text(
                    "⚠️ Неверный формат даты. Используйте формат ДД.ММ.ГГГГ, например: 25.03.2024",
                    reply_markup=create_notes_menu_keyboard()
                )
                return
        else:
            result = search_notes(user_id, query=text)

        await update.message.reply_text(
            result,
            reply_markup=create_notes_menu_keyboard()
        )
        return

    # Остальной код обработки текста...
    if user_id in user_calendar_date:
        selected_date = user_calendar_date[user_id]
        save_note(user_id, selected_date, text)
        del user_calendar_date[user_id]

        await update.message.reply_text(
            f"✅ Заметка на {selected_date.strftime('%d.%m.%Y')} сохранена!",
            reply_markup=create_main_menu_keyboard(user_id)
        )
        return

    # Обработка группы
    group = text.upper()
    if group not in available_groups:
        await update.message.reply_text(
            "❌ Такой группы нет. Введите корректное название.",
            reply_markup=create_main_menu_keyboard(user_id)
        )
        return

    user_groups[user_id] = group
    await show_schedule_with_day_selector(update, context, user_id, group)


def create_direction_keyboard():
    directions = sorted(set(group[:3] for group in available_groups))
    buttons = [[InlineKeyboardButton(direction, callback_data=f"direction_{direction}")] for direction in directions]
    buttons.append([InlineKeyboardButton("◀ Назад", callback_data="go_back_to_main")])
    return InlineKeyboardMarkup(buttons)


def create_groups_keyboard(direction):
    groups = sorted([group for group in available_groups if group.startswith(direction)])
    buttons = [[InlineKeyboardButton(group, callback_data=f"group_{group}")] for group in groups]
    buttons.append([InlineKeyboardButton("◀ К направлениям", callback_data="back_to_directions")])
    return InlineKeyboardMarkup(buttons)


def create_day_selector_keyboard():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("Пн", callback_data="day_mon"),
            InlineKeyboardButton("Вт", callback_data="day_tue"),
            InlineKeyboardButton("Ср", callback_data="day_wed"),
            InlineKeyboardButton("Чт", callback_data="day_thu")
        ],
        [
            InlineKeyboardButton("Пт", callback_data="day_fri"),
            InlineKeyboardButton("Сб", callback_data="day_sat"),
            InlineKeyboardButton("Вс", callback_data="day_sun"),
            InlineKeyboardButton("✔ Сегодня", callback_data="day_today")
        ],
        [
            InlineKeyboardButton("◀️ Пред. неделя", callback_data="prev_week"),
            InlineKeyboardButton("След. неделя ▶️", callback_data="next_week")
        ],
        [InlineKeyboardButton("◀ Назад", callback_data="go_back_to_main")]
    ])


async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    if query.data == "admin_panel":
        if user_id in ADMINS:
            await query.edit_message_text(
                "Панель администратора:",
                reply_markup=create_admin_keyboard()
            )
        return

    elif query.data == "notes_menu":
        await query.edit_message_text(
            "🗂 Управление заметками:",
            reply_markup=create_notes_menu_keyboard()
        )
        return

    elif query.data == "show_all":
        result = search_notes(user_id)
        keyboard = [[InlineKeyboardButton("◀ Назад", callback_data="notes_menu")]]
        await query.edit_message_text(
            result,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return

    elif query.data in ["search_text", "search_date"]:
        context.user_data['search_type'] = query.data
        keyboard = [[InlineKeyboardButton("◀ Отмена", callback_data="notes_menu")]]
        if query.data == "search_text":
            await query.edit_message_text(
                "Введите текст для поиска по заметкам:",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
        else:
            await query.edit_message_text(
                "Введите дату для поиска в формате ДД.ММ.ГГГГ:",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
        context.user_data['waiting_for_search'] = True
        return

    elif query.data == "upload_schedule":
        return await request_schedule_file(update, context)

    elif query.data == "cancel_upload":
        await query.edit_message_text(
            "Загрузка отменена.",
            reply_markup=create_main_menu_keyboard(user_id)
        )
        return ConversationHandler.END

    elif query.data == "show_calendar":
        today = datetime.datetime.now()
        await query.edit_message_text(
            "Выберите дату для заметки:",
            reply_markup=create_calendar_keyboard(today.year, today.month)
        )
        return

    elif query.data.startswith("month_"):
        _, year, month = query.data.split("_")
        await query.edit_message_text(
            "Выберите дату для заметки:",
            reply_markup=create_calendar_keyboard(int(year), int(month))
        )
        return

    elif query.data.startswith("calendar_"):
        _, year, month, day = query.data.split("_")
        selected_date = datetime.datetime(int(year), int(month), int(day))
        user_calendar_date[user_id] = selected_date

        await query.edit_message_text(
            f"Введите заметку для даты {selected_date.strftime('%d.%m.%Y')}:",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀ Отмена", callback_data="cancel_note")
            ]])
        )
        return WAITING_FOR_NOTE

    elif query.data == "cancel_note" or query.data == "cancel_calendar":
        await query.edit_message_text(
            "Выберите действие:",
            reply_markup=create_main_menu_keyboard(user_id)
        )
        return

    elif query.data == "view_schedule":
        if not available_groups:
            if user_id in ADMINS:
                await query.edit_message_text(
                    "Расписание ещё не загружено. Загрузите его через админ-панель.",
                    reply_markup=create_main_menu_keyboard(user_id)
                )
            else:
                await query.edit_message_text(
                    "Расписание ещё не загружено. Обратитесь к администратору.",
                    reply_markup=create_main_menu_keyboard(user_id)
                )
        else:
            group = user_groups.get(user_id)
            if not group:
                keyboard = create_initial_group_keyboard()
                await query.edit_message_text(
                    "Сначала выберите группу:",
                    reply_markup=keyboard
                )
            else:
                current_day_selections[user_id] = datetime.datetime.now()
                await show_schedule_with_day_selector(update, context, user_id, group)
        return

    elif query.data == "back_to_directions":
        await query.edit_message_text(
            "Выберите направление:",
            reply_markup=create_direction_keyboard()
        )
        return

    elif query.data == "go_back_to_main":
        await query.edit_message_text(
            "Выберите действие:",
            reply_markup=create_main_menu_keyboard(user_id)
        )
        return

    elif query.data.startswith("direction_"):
        direction = query.data.split("_")[1]
        await query.edit_message_text(
            f"Выберите группу направления {direction}:",
            reply_markup=create_groups_keyboard(direction)
        )
        return

    elif query.data.startswith("group_"):
        group = query.data.split("_")[1]
        user_groups[user_id] = group
        current_day_selections[user_id] = datetime.datetime.now()
        await show_schedule_with_day_selector(update, context, user_id, group)
        return

    elif query.data == "day_today":
        group = user_groups.get(user_id, "")
        today = datetime.datetime.now()
        current_day_selections[user_id] = today
        await show_schedule_with_day_selector(update, context, user_id, group, today)
        return

    elif query.data.startswith("day_"):
        day_map = {
            "mon": 0, "tue": 1, "wed": 2,
            "thu": 3, "fri": 4, "sat": 5, "sun": 6
        }
        day_key = query.data.split("_")[1]
        if day_key in day_map:
            current_date = current_day_selections.get(user_id, datetime.datetime.now())
            monday_of_week = current_date - datetime.timedelta(days=current_date.weekday())
            selected_date = monday_of_week + datetime.timedelta(days=day_map[day_key])
            current_day_selections[user_id] = selected_date

            group = user_groups.get(user_id, "")
            await show_schedule_with_day_selector(update, context, user_id, group, selected_date)
        return

    elif query.data in ["prev_week", "next_week"]:
        group = user_groups.get(user_id, "")
        if not group:
            await query.edit_message_text(
                "Сначала выберите группу.",
                reply_markup=create_main_menu_keyboard(user_id)
            )
            return

        current_date = current_day_selections.get(user_id, datetime.datetime.now())
        monday_of_week = current_date - datetime.timedelta(days=current_date.weekday())
        new_date = monday_of_week + datetime.timedelta(weeks=(-1 if query.data == "prev_week" else 1))
        new_date = new_date + datetime.timedelta(days=current_date.weekday())
        current_day_selections[user_id] = new_date

        await show_schedule_with_day_selector(update, context, user_id, group, new_date)
        return

    elif query.data == "toggle_subscription":
        if user_id in subscribed_users:
            subscribed_users.remove(user_id)
            await query.answer("Вы отключили рассылку.")
        else:
            subscribed_users.add(user_id)
            await query.answer("Вы включили рассылку.")
        # Обновляем главное меню с новым статусом рассылки
        await query.edit_message_text(
            "Выберите действие:",
            reply_markup=create_main_menu_keyboard(user_id)
        )
        return

    elif query.data == "delete_note":
        if user_id not in user_notes or not user_notes[user_id]:
            await query.edit_message_text(
                "У вас пока нет сохраненных заметок.",
                reply_markup=create_notes_menu_keyboard()
            )
            return

        context.user_data['waiting_for_delete'] = True
        keyboard = [[InlineKeyboardButton("◀ Отмена", callback_data="notes_menu")]]
        await query.edit_message_text(
            "Введите дату заметки для удаления в формате ДД.ММ.ГГГГ:\n"
            "Например: 25.03.2024",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return

    elif query.data == "ai_assistant":
        keyboard = [[InlineKeyboardButton("◀ Назад", callback_data="go_back_to_main")]]
        message = await query.edit_message_text(
            "❓ Я ваш помощник! Вы можете спросить меня:\n\n"
            "• Когда следующая [предмет]?\n"
            "• Сколько осталось пар на этой неделе?\n\n"
            "Просто напишите свой вопрос в чат.",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        context.user_data['last_assistant_message'] = message.message_id
        return

    elif query.data.startswith("course_"):
        course = int(query.data.split("_")[1])
        context.user_data['selected_course'] = course
        await query.edit_message_text(
            f"Выберите направление ({course} курс):",
            reply_markup=create_direction_keyboard_for_course(course)
        )
        return

    elif query.data == "back_to_courses":
        keyboard = create_course_keyboard()
        await query.edit_message_text(
            "Выберите курс:",
            reply_markup=keyboard
        )
        return

    elif query.data.startswith("initial_direction_"):
        direction = query.data.replace("initial_direction_", "")
        course = context.user_data.get('selected_course')
        if course:
            await query.edit_message_text(
                f"Выберите группу направления {direction} ({course} курс):",
                reply_markup=create_groups_keyboard_for_direction(direction, course)
            )
        return

    elif query.data.startswith("back_to_directions_"):
        course = int(query.data.split("_")[-1])
        context.user_data['selected_course'] = course
        await query.edit_message_text(
            f"Выберите направление ({course} курс):",
            reply_markup=create_direction_keyboard_for_course(course)
        )
        return

    elif query.data.startswith("initial_group_"):
        group = query.data.replace("initial_group_", "")
        user_groups[user_id] = group
        await query.edit_message_text(
            f"✅ Группа {group} выбрана!\n\nВыберите действие:",
            reply_markup=create_main_menu_keyboard(user_id)
        )
        return

    elif query.data == "change_group":
        keyboard = create_course_keyboard()
        await query.edit_message_text(
            "Выберите курс:",
            reply_markup=keyboard
        )
        return

    elif query.data == "set_time":
        await set_notification_time(update, context)
        return

    elif query.data.startswith("time_"):
        hour = int(query.data.split("_")[1])
        user_id = query.from_user.id
        user_notification_times[user_id] = hour

        # Обновляем расписание в планировщике для этого пользователя
        scheduler = context.job_queue
        user_jobs = scheduler.get_jobs_by_name(f"daily_notification_{user_id}")
        for job in user_jobs:
            job.schedule_removal()

        # Создаем новое расписание для пользователя
        scheduler.run_daily(
            send_daily_schedule,
            time=datetime.time(hour=hour, minute=0),
            name=f"daily_notification_{user_id}"
        )

        await query.edit_message_text(
            f"✅ Время рассылки установлено на {hour:02d}:00",
            reply_markup=create_main_menu_keyboard(user_id)
        )
        return


async def show_schedule_with_day_selector(update, context, user_id, group, date=None):
    if date is None:
        date = datetime.datetime.now()

    schedule = get_schedule_by_date(group, date)
    keyboard = create_day_selector_keyboard()

    if isinstance(update, Update):
        if update.message:
            await update.message.reply_text(schedule, parse_mode='HTML', reply_markup=keyboard)
        elif update.callback_query:
            await update.callback_query.edit_message_text(schedule, parse_mode='HTML', reply_markup=keyboard)
    else:
        await update.edit_message_text(schedule, parse_mode='HTML', reply_markup=keyboard)


def get_weather_forecast(city=CITY_NAME):
    try:
        # Получаем прогноз на день
        url = f"http://api.openweathermap.org/data/2.5/forecast?q={city}&appid={WEATHER_API_KEY}&units=metric&lang=ru"
        response = requests.get(url)
        data = response.json()

        if response.status_code == 200:
            # Получаем прогноз на следующий день
            tomorrow_forecasts = []
            tomorrow = datetime.datetime.now() + datetime.timedelta(days=1)
            tomorrow_date = tomorrow.strftime('%Y-%m-%d')

            for item in data['list']:
                forecast_time = datetime.datetime.fromtimestamp(item['dt'])
                if forecast_time.strftime('%Y-%m-%d') == tomorrow_date:
                    hour = forecast_time.hour
                    if hour in [12, 18]:  # Берем прогноз на день (12:00) и вечер (18:00)
                        tomorrow_forecasts.append({
                            'temp': item['main']['temp'],
                            'description': item['weather'][0]['description'],
                            'time': 'днем' if hour == 12 else 'ближе к вечеру'
                        })

            if tomorrow_forecasts:
                forecast_text = "👋 Добрый вечер!\n\nЗавтра "
                for forecast in tomorrow_forecasts:
                    forecast_text += f"{forecast['time']}😬 🌡️ {forecast['temp']:.1f}°C, {forecast['description']}, "
                return forecast_text.rstrip(', ') + ".\n\n"
        return "👋 Добрый вечер!\n\n"
    except Exception as e:
        print(f"Ошибка получения прогноза погоды: {e}")
        return "👋 Добрый вечер!\n\n"


def save_note(user_id: int, date: datetime.datetime, note: str):
    date_str = date.strftime("%Y-%m-%d")
    if user_id not in user_notes:
        user_notes[user_id] = {}
    user_notes[user_id][date_str] = note.strip()


def get_note(user_id: int, date: datetime.datetime) -> str:
    date_str = date.strftime("%Y-%m-%d")
    return user_notes.get(user_id, {}).get(date_str, "")


async def add_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if user_id not in user_groups:
        await update.message.reply_text("Сначала выберите группу!")
        return ConversationHandler.END

    date = current_day_selections.get(user_id, datetime.datetime.now())
    date_str = date.strftime("%Y-%m-%d")
    context.user_data['note_date'] = date_str

    await update.message.reply_text(
        f"Введите заметку для {date_str}:\n"
        "Для отмены введите /cancel"
    )
    return WAITING_FOR_NOTE


async def save_note_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    note_text = update.message.text
    date_str = context.user_data.get('note_date')

    if date_str:
        date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        save_note(user_id, date, note_text)
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("◀ Назад", callback_data="go_back_to_main")
        ]])
        formatted_date = format_date_russian(date)
        await update.message.reply_text(
            f"✅ Заметка на {formatted_date} сохранена!",
            reply_markup=keyboard
        )

    return ConversationHandler.END


async def cancel_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Добавление заметки отменено.")
    return ConversationHandler.END


async def send_daily_schedule(context: ContextTypes.DEFAULT_TYPE, user_ids=None):
    tomorrow = datetime.datetime.now() + datetime.timedelta(days=1)
    weather = get_weather_forecast()

    # Если user_ids не указан, отправляем всем подписанным пользователям
    target_users = user_ids if user_ids else subscribed_users

    for user_id in target_users:
        if user_id not in user_groups:
            continue

        group = user_groups[user_id]
        schedule = get_schedule_by_date(group, tomorrow)
        note = get_note(user_id, tomorrow)

        message = weather
        message += "Пары на завтра:\n"
        message += schedule

        if note:
            message += f"\n❗❗❗Ты хотел не забыть про это:\n{note}"

        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("◀ Назад", callback_data="go_back_to_main")
        ]])

        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=message,
                parse_mode='HTML',
                reply_markup=keyboard
            )
        except Exception as e:
            print(f"Error sending message to user {user_id}: {e}")


async def test_notification(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if user_id not in ADMINS:
        await update.message.reply_text("У вас нет доступа к этой команде.")
        return

    await update.message.reply_text("Отправляю тестовую рассылку...")
    await send_daily_schedule(context)
    await update.message.reply_text("Тестовая рассылка отправлена.")


def format_date_russian(date):
    return date.strftime("%d %B %Y").replace("January", "января").replace("February", "февраля") \
        .replace("March", "марта").replace("April", "апреля").replace("May", "мая") \
        .replace("June", "июня").replace("July", "июля").replace("August", "августа") \
        .replace("September", "сентября").replace("October", "октября") \
        .replace("November", "ноября").replace("December", "декабря")


def search_notes(user_id: int, query: str = None, date_str: str = None) -> str:
    if user_id not in user_notes or not user_notes[user_id]:
        return "У вас пока нет сохраненных заметок. Чтобы добавить заметку, вернитесь в главное меню и нажмите '📝 Добавить заметку'"

    found_notes = []

    # Если нет параметров поиска, показываем все заметки
    if not query and not date_str:
        found_notes = [(date, text) for date, text in user_notes[user_id].items()]
    else:
        # Поиск по конкретным параметрам
        for note_date, note_text in user_notes[user_id].items():
            # Поиск по дате
            if date_str and date_str in note_date:
                found_notes.append((note_date, note_text))
                continue

            # Поиск по тексту
            if query and query.lower() in note_text.lower():
                found_notes.append((note_date, note_text))

    if not found_notes:
        if query or date_str:
            search_term = f"дате {date_str}" if date_str else f"запросу '{query}'"
            return f"🔍 Заметки по {search_term} не найдены."
        else:
            return "У вас пока нет сохраненных заметок. Чтобы добавить заметку, вернитесь в главное меню и нажмите '📝 Добавить заметку'"

    # Сортируем заметки по дате (сначала новые)
    found_notes.sort(key=lambda x: x[0], reverse=True)

    if not query and not date_str:
        result = "📝 Ваши заметки:\n\n"
    else:
        search_term = f"дате {date_str}" if date_str else f"запросу '{query}'"
        result = f"🔍 Результаты поиска по {search_term}:\n\n"

    # Форматируем каждую заметку
    for note_date, note_text in found_notes:
        date_obj = datetime.datetime.strptime(note_date, "%Y-%m-%d")
        formatted_date = format_date_russian(date_obj)
        result += f"📅 {formatted_date}:\n{note_text}\n\n"

    # Добавляем информацию о количестве найденных заметок
    total_notes = len(found_notes)
    result += f"\nВсего заметок: {total_notes}"

    return result.strip()


async def show_notes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    result = search_notes(user_id)
    await update.message.reply_text(result)


async def start_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("🔍 Поиск по тексту", callback_data="search_text")],
        [InlineKeyboardButton("📅 Поиск по дате", callback_data="search_date")],
        [InlineKeyboardButton("📝 Показать все заметки", callback_data="show_all")],
        [InlineKeyboardButton("◀ Отмена", callback_data="cancel_search")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        "Выберите тип поиска:",
        reply_markup=reply_markup
    )


async def handle_search_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    if query.data == "show_all":
        result = search_notes(user_id)
        await query.edit_message_text(result)
        return ConversationHandler.END

    elif query.data == "cancel_search":
        await query.edit_message_text("Поиск отменен.")
        return ConversationHandler.END

    elif query.data in ["search_text", "search_date"]:
        context.user_data['search_type'] = query.data
        if query.data == "search_text":
            await query.edit_message_text(
                "Введите текст для поиска по заметкам:\n"
                "(для отмены введите /cancel)"
            )
        else:
            await query.edit_message_text(
                "Введите дату для поиска в формате ДД.ММ.ГГГГ:\n"
                "(для отмены введите /cancel)"
            )
        return WAITING_FOR_SEARCH


async def handle_search_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    search_input = update.message.text

    if search_input == '/cancel':
        await update.message.reply_text("Поиск отменен.")
        return ConversationHandler.END

    search_type = context.user_data.get('search_type')

    if search_type == "search_date":
        try:
            # Преобразуем введенную дату в формат хранения
            date_obj = datetime.datetime.strptime(search_input, "%d.%m.%Y")
            search_date = date_obj.strftime("%Y-%m-%d")
            result = search_notes(user_id, date_str=search_date)
        except ValueError:
            await update.message.reply_text(
                "Неверный формат даты. Используйте формат ДД.ММ.ГГГГ, например: 25.03.2024"
            )
            return WAITING_FOR_SEARCH
    else:
        result = search_notes(user_id, query=search_input)

    await update.message.reply_text(result)
    return ConversationHandler.END


def delete_note(user_id: int, date: datetime.datetime) -> bool:
    date_str = date.strftime("%Y-%m-%d")
    if user_id in user_notes and date_str in user_notes[user_id]:
        del user_notes[user_id][date_str]
        return True
    return False


def find_next_lesson(user_id: int, subject_query: str) -> str:
    if user_id not in user_groups:
        return "❌ Сначала выберите группу в главном меню!"

    group = user_groups[user_id]
    today = datetime.datetime.now()

    # Ищем похожие предметы
    subject_query = subject_query.lower()
    matching_subjects = set()
    for item in schedule_data:
        if item['group'] == group and subject_query in item['subject'].lower():
            matching_subjects.add(item['subject'])

    if not matching_subjects:
        return f"❌ Предмет, похожий на '{subject_query}', не найден в вашем расписании."

    # Для каждого похожего предмета ищем ближайшую пару
    next_lessons = []
    for subject in matching_subjects:
        # Проверяем на неделю вперед
        for i in range(8):  # сегодня + 7 дней
            check_date = today + datetime.timedelta(days=i)
            week_num = get_current_week(check_date)
            weekday = check_date.strftime("%A")
            russian_days = {
                "Monday": "Понедельник", "Tuesday": "Вторник", "Wednesday": "Среда",
                "Thursday": "Четверг", "Friday": "Пятница", "Saturday": "Суббота",
                "Sunday": "Воскресенье"
            }
            day_rus = russian_days.get(weekday)

            # Ищем пары этого предмета в этот день
            for item in schedule_data:
                if (item['group'] == group and
                        item['subject'] == subject and
                        item['day'] == day_rus and
                        (item['week_num'] is None or item['week_num'] == week_num)):
                    next_lessons.append({
                        'subject': subject,
                        'date': check_date,
                        'time': item['time'],
                        'type': item['type'],
                        'room': item['room'],
                        'teacher': item['teacher'],
                        'days_until': i
                    })

    if not next_lessons:
        return f"❌ Не нашел предстоящих занятий по предмету '{subject_query}' в ближайшую неделю."

    # Сортируем по дате и времени
    next_lessons.sort(key=lambda x: (x['days_until'], x['time']))
    next_lesson = next_lessons[0]

    # Формируем ответ
    if next_lesson['days_until'] == 0:
        day_str = "сегодня"
    elif next_lesson['days_until'] == 1:
        day_str = "завтра"
    else:
        day_str = format_date_russian(next_lesson['date'])

    return (f"📚 Следующее занятие по предмету '{next_lesson['subject']}':\n"
            f"📅 {day_str}, {next_lesson['time']}\n"
            f"📝 {next_lesson['type']}\n"
            f"🏛 Аудитория: {next_lesson['room']}\n"
            f"👨‍🏫 Преподаватель: {next_lesson['teacher']}")


def count_remaining_lessons(user_id: int) -> str:
    if user_id not in user_groups:
        return "❌ Сначала выберите группу в главном меню!"

    group = user_groups[user_id]
    today = datetime.datetime.now()
    current_week_num = get_current_week(today)

    # Считаем оставшиеся пары на этой неделе
    remaining_lessons = []
    for i in range(8):  # проверяем текущую неделю
        check_date = today + datetime.timedelta(days=i)
        if check_date.isocalendar()[1] != today.isocalendar()[1]:  # если перешли на следующую неделю
            break

        weekday = check_date.strftime("%A")
        russian_days = {
            "Monday": "Понедельник", "Tuesday": "Вторник", "Wednesday": "Среда",
            "Thursday": "Четверг", "Friday": "Пятница", "Saturday": "Суббота",
            "Sunday": "Воскресенье"
        }
        day_rus = russian_days.get(weekday)

        # Если это сегодня, учитываем текущее время
        current_time = None
        if i == 0:
            current_time = today.strftime("%H:%M")

        for item in schedule_data:
            if (item['group'] == group and
                    item['day'] == day_rus and
                    (item['week_num'] is None or item['week_num'] == current_week_num)):

                # Если это сегодня, проверяем, не прошла ли уже пара
                if current_time:
                    lesson_end_time = item['time'].split('-')[1]
                    if current_time > lesson_end_time:
                        continue

                remaining_lessons.append({
                    'subject': item['subject'],
                    'date': check_date,
                    'time': item['time'],
                    'type': item['type']
                })

    if not remaining_lessons:
        return "🎉 На этой неделе больше нет пар!"

    # Сортируем по дате и времени
    remaining_lessons.sort(key=lambda x: (x['date'], x['time']))

    # Формируем ответ
    total = len(remaining_lessons)
    response = f"📚 Осталось {total} пар на этой неделе:\n\n"

    current_date = None
    for lesson in remaining_lessons:
        if current_date != lesson['date']:
            current_date = lesson['date']
            if lesson['date'].date() == today.date():
                response += f"\n📅 Сегодня:\n"
            elif lesson['date'].date() == (today + datetime.timedelta(days=1)).date():
                response += f"\n📅 Завтра:\n"
            else:
                response += f"\n📅 {lesson['date'].strftime('%d.%m.%Y')}:\n"

        response += f"⏰ {lesson['time']} - {lesson['subject']} ({lesson['type']})\n"

    return response


async def handle_ai_request(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    text = update.message.text.lower()

    # Проверяем, выбрана ли группа
    if user_id not in user_groups:
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("Выбрать группу", callback_data="view_schedule"),
            InlineKeyboardButton("◀ Назад", callback_data="go_back_to_main")
        ]])
        message = await update.message.reply_text(
            "❌ Сначала нужно выбрать группу!\n"
            "Нажмите 'Выбрать группу' чтобы продолжить.",
            reply_markup=keyboard
        )
        context.user_data['last_assistant_message'] = message.message_id
        return

    # Удаляем предыдущее сообщение помощника, если оно есть
    if 'last_assistant_message' in context.user_data:
        try:
            await context.bot.delete_message(
                chat_id=user_id,
                message_id=context.user_data['last_assistant_message']
            )
        except Exception:
            pass  # Игнорируем ошибки при удалении сообщения

    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("◀ Назад", callback_data="go_back_to_main")
    ]])

    # Проверяем различные типы запросов
    if "когда" in text and "следующ" in text:
        # Ищем название предмета в запросе
        subject_words = text.split()
        for word in ["когда", "следующ", "следующая", "следующий", "будет"]:
            if word in subject_words:
                subject_words.remove(word)
        subject_query = " ".join(subject_words).strip()

        response = find_next_lesson(user_id, subject_query)
        message = await update.message.reply_text(response, reply_markup=keyboard)
        context.user_data['last_assistant_message'] = message.message_id
        return

    elif "сколько" in text and ("пар" in text or "занятий" in text) and "недел" in text:
        response = count_remaining_lessons(user_id)
        message = await update.message.reply_text(response, reply_markup=keyboard)
        context.user_data['last_assistant_message'] = message.message_id
        return

    # Если запрос не распознан
    message = await update.message.reply_text(
        "🤔 Я пока не умею отвечать на такой вопрос. Попробуйте спросить:\n"
        "• Когда следующая [предмет]?\n"
        "• Сколько осталось пар на этой неделе?",
        reply_markup=keyboard
    )
    context.user_data['last_assistant_message'] = message.message_id


def create_initial_group_keyboard():
    if not available_groups:
        return None

    # Получаем уникальные направления
    directions = sorted(set(group[:3] for group in available_groups))

    keyboard = []
    for direction in directions:
        keyboard.append([InlineKeyboardButton(direction, callback_data=f"initial_direction_{direction}")])

    if len(keyboard) == 0:
        return None

    return InlineKeyboardMarkup(keyboard)


def create_initial_groups_keyboard(direction):
    groups = sorted([group for group in available_groups if group.startswith(direction)])
    buttons = [[InlineKeyboardButton(group, callback_data=f"initial_group_{group}")] for group in groups]
    buttons.append([InlineKeyboardButton("◀ К направлениям", callback_data="back_to_initial_directions")])
    return InlineKeyboardMarkup(buttons)


async def set_notification_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("◀ Отмена", callback_data="go_back_to_main")
    ]])

    await query.edit_message_text(
        "⏰ Введите желаемое время для ежедневной рассылки в формате ЧЧ:ММ\n"
        "Например: 19:30 или 20:00",
        reply_markup=keyboard
    )
    return WAITING_FOR_TIME


async def handle_time_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    text = update.message.text.strip()

    try:
        # Пробуем разобрать введенное время
        if ':' in text:
            hours, minutes = map(int, text.split(':'))
        else:
            hours = int(text)
            minutes = 0

        # Проверяем корректность времени
        if not (0 <= hours <= 23 and 0 <= minutes <= 59):
            raise ValueError

        # Сохраняем время пользователя
        user_notification_times[user_id] = hours

        # Обновляем расписание в планировщике для этого пользователя
        scheduler = context.job_queue
        user_jobs = scheduler.get_jobs_by_name(f"daily_notification_{user_id}")
        for job in user_jobs:
            job.schedule_removal()

        # Создаем новое расписание для пользователя
        scheduler.run_daily(
            send_daily_schedule,
            time=datetime.time(hour=hours, minute=minutes),
            name=f"daily_notification_{user_id}"
        )

        # Отправляем подтверждение и возвращаем в главное меню
        await update.message.reply_text(
            f"✅ Время рассылки установлено на {hours:02d}:{minutes:02d}",
            reply_markup=create_main_menu_keyboard(user_id)
        )
        return ConversationHandler.END

    except (ValueError, IndexError):
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("◀ Отмена", callback_data="go_back_to_main")
        ]])
        await update.message.reply_text(
            "❌ Неверный формат времени. Пожалуйста, используйте формат ЧЧ:ММ\n"
            "Например: 19:30 или 20:00",
            reply_markup=keyboard
        )
        return WAITING_FOR_TIME


def get_user_notification_time(user_id: int) -> int:
    return user_notification_times.get(user_id, DEFAULT_NOTIFICATION_TIME)


def main():
    load_schedule()
    application = ApplicationBuilder().token(TOKEN).build()

    # Настройка планировщика для ежедневной рассылки
    scheduler = BackgroundScheduler(timezone=pytz.timezone('Europe/Moscow'))

    # Создаем отдельные задачи для каждого пользователя
    for user_id in subscribed_users:
        notification_time = get_user_notification_time(user_id)
        scheduler.add_job(
            lambda uid=user_id: application.job_queue.run_once(
                lambda ctx: send_daily_schedule(ctx, user_ids=[uid]), 0
            ),
            'cron',
            hour=notification_time,
            minute=0
        )
    scheduler.start()

    # Создаем ConversationHandler для установки времени
    time_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(set_notification_time, pattern="^set_time$")],
        states={
            WAITING_FOR_TIME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_time_input)
            ]
        },
        fallbacks=[CallbackQueryHandler(lambda u, c: ConversationHandler.END, pattern="^go_back_to_main$")],
        per_chat=True,
        per_message=False
    )

    # Создаем ConversationHandler для поиска
    search_handler = ConversationHandler(
        entry_points=[CommandHandler('search', start_search)],
        states={
            WAITING_FOR_SEARCH: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_search_input)
            ]
        },
        fallbacks=[CommandHandler('cancel', lambda u, c: ConversationHandler.END)],
        per_chat=True,
        per_message=False
    )

    # Создаем ConversationHandler для загрузки файла
    file_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(request_schedule_file, pattern="^upload_schedule$")],
        states={
            WAITING_FOR_FILE: [
                MessageHandler(filters.Document.ALL, handle_document)
            ]
        },
        fallbacks=[CallbackQueryHandler(lambda u, c: ConversationHandler.END, pattern="^cancel_upload$")],
        per_chat=True,
        per_message=False
    )

    # Обработчики команд и сообщений
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("test", test_notification))
    application.add_handler(search_handler)
    application.add_handler(file_handler)
    application.add_handler(time_handler)

    # Добавляем общий обработчик callback_query до обработчика текстовых сообщений
    application.add_handler(CallbackQueryHandler(button_callback))

    # Обработчик текстовых сообщений должен быть последним
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    print("Бот запущен...")
    application.run_polling()


if __name__ == '__main__':
    main()
